import telebot
from telebot import types
import datetime
import schedule as shl
import time
import threading
import openpyxl

file_name_stat = r"C:\Users\allro\Documents\Python Notebooks\antidepression_stat.xlsx"  # Файл для хранения настроений
file_name_schedule = r"C:\Users\allro\Documents\Python Notebooks\antidepression_schedule.xlsx"  # Файл для хранения времени напоминания
file_name_beck = r"C:\Users\allro\Documents\Python Notebooks\antidepression_beck.xlsx"  # Файл для хранения результатов теста Бека
token_location = r"C:\Users\allro\Documents\Python Notebooks\tokenapi.txt"  # В этот файл кладем токен бота

# Получение токена бота
with open(token_location, 'r', encoding='utf-8') as file:
    API_TOKEN = file.read()

bot = telebot.TeleBot(API_TOKEN)

# Храним чат-айди пользователей и их имена
user_names = {}

# Все вопросы из теста Бека
questions = [
    '1. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я не чувствую себя несчастным.\n   1. Я чувствую себя несчастным.\n   2. Я всё время несчастен и не могу освободиться от этого чувства.\n   3. Я настолько несчастен и опечален, что не могу этого вынести.',
    '2. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Думая о будущем, я не чувствую себя особенно разочарованным.\n   1. Думая о будущем, я чувствую себя разочарованным.\n   2. Я чувствую, что мне нечего ждать в будущем.\n   3. Я чувствую, что будущее безнадежно и ничто не изменится к лучшему.',
    '3. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я не чувствую себя неудачником.\n   1. Я чувствую, что у меня было больше неудач, чем у большинства других людей.\n   2. Когда я оглядываюсь на прожитую жизнь, всё, что я вижу, это череда неудач.\n   3. Я чувствую себя полным неудачником.',
    '4. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я получаю столько же удовольствия от жизни, как и раньше.\n   1. Я не получаю столько же удовольствия от жизни, как раньше.\n   2. Я не получаю настоящего удовлетворения от чего бы то ни было.\n   3. Я всем не удовлетворен, и мне всё надоело.',
    '5. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я не чувствую себя особенно виноватым.\n   1. Довольно часто я чувствую себя виноватым.\n   2. Почти всегда я чувствую себя виноватым.\n   3. Я чувствую себя виноватым всё время.',
    '6. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я не чувствую, что меня за что-то наказывают.\n   1. Я чувствую, что могу быть наказан за что-то.\n   2. Я ожидаю, что меня накажут.\n   3. Я чувствую, что меня наказывают за что-то.',
    '7. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я не испытываю разочарования в себе.\n   1. Я разочарован в себе.\n   2. Я внушаю себе отвращение.\n   3. Я ненавижу себя.',
    '8. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. У меня нет чувства, что я в чем-то хуже других.\n   1. Я самокритичен и признаю свои слабости и ошибки.\n   2. Я всё время виню себя за свои ошибки.\n   3. Я виню себя за всё плохое, что происходит.',
    '9. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. У меня нет мыслей о том, чтобы покончить с собой.\n   1. У меня есть мысли о том, чтобы покончить с собой, но я этого не сделаю.\n   2. Я хотел бы покончить жизнь самоубийством.\n   3. Я бы покончил с собой, если бы представился удобный случай.',
    '10. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я плачу не больше, чем обычно.\n   1. Сейчас я плачу больше обычного.\n   2. Я теперь всё время плачу.\n   3. Раньше я еще мог плакать, но теперь не смогу, даже если захочу.',
    '11. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Сейчас я не более раздражен, чем обычно.\n   1. Я раздражаюсь легче, чем раньше, даже по пустякам.\n   2. Сейчас я всё время раздражен.\n   3. Меня уже ничто не раздражает, потому что всё стало безразлично.',
    '12. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я не потерял интереса к другим людям.\n   1. У меня меньше интереса к другим людям, чем раньше.\n   2. Я почти утратил интерес к другим людям.\n   3. Я потерял всякий интерес к другим людям.',
    '13. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я способен принимать решения так же, как всегда.\n   1. Я откладываю принятие решений чаще, чем обычно.\n   2. Я испытываю больше трудностей в принятии решений, чем прежде.\n   3. Я больше не могу принимать каких-либо решений.',
    '14. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я не чувствую, что я выгляжу хуже, чем обычно.\n   1. Я обеспокоен тем, что выгляжу постаревшим или непривлекательным.\n   2. Я чувствую, что изменения, происшедшие в моей внешности, сделали меня непривлекательным.\n   3. Я уверен, что выгляжу безобразным.',
    '15. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я могу работать так же, как раньше.\n   1. Мне надо приложить дополнительные усилия, чтобы начать что-либо делать.\n   2. Я с большим трудом заставляю себя что-либо делать.\n   3. Я вообще не могу работать.',
    '16. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я могу спать так же хорошо, как и обычно.\n   1. Я сплю не так хорошо, как всегда.\n   2. Я просыпаюсь на 1-2 часа раньше, чем обычно, и с трудом могу заснуть снова.\n   3. Я просыпаюсь на несколько часов раньше обычного и не могу снова заснуть.',
    '17. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я устаю не больше обычного.\n   1. Я устаю легче обычного.\n   2. Я устаю почти от всего того, что я делаю.\n   3. Я слишком устал, чтобы делать что бы то ни было.',
    '18. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Мой аппетит не хуже, чем обычно.\n   1. У меня не такой хороший аппетит, как был раньше.\n   2. Сейчас мой аппетит стал намного хуже.\n   3. Я вообще потерял аппетит.',
    '19. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Если в последнее время я и потерял в весе, то очень немного.\n   1. Я потерял в весе более 2 кг.\n   2. Я потерял в весе более 4 кг.\n   3. Я потерял в весе более 6 кг.',
    '20. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я беспокоюсь о своем здоровье не больше, чем обычно.\n   1. Меня беспокоят такие проблемы, как различные боли, расстройства желудка, запоры.\n   2. Я настолько обеспокоен своим здоровьем, что мне даже трудно думать о чем-нибудь другом.\n   3. Я до такой степени обеспокоен своим здоровьем, что вообще ни о чем другом не могу думать.',
    '21. Что лучше описывает ваше состояние за прошедшую неделю и сегодня?\n   0. Я не замечал каких-либо изменений в моих сексуальных интересах.\n   1. Я меньше, чем обычно интересуюсь сексом.\n   2. Сейчас я намного меньше интересуюсь сексом.\n   3. Я совершенно утратил интерес к сексу.'
]


def write_stat_to_file(file_name: str, user_name: str, mark: int, record_time_string: str) -> None:
    """Записывает информацию по настроению пользователя в файл"""
    # Проверяем, существует ли файл
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        # Если файл не существует, создаем новый и добавляем заголовки
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["user_name", "mark", "record_time_string"])  # Заголовки столбцов

    # Флаг для отслеживания наличия пользователя
    user_exists = False

    # Проверяем существующие записи
    for row in sheet.iter_rows(min_row=2):  # Пропускаем заголовок
        if row[0].value == user_name and row[2].value == record_time_string:
            row[1].value = mark  # Обновляем время для существующего пользователя
            user_exists = True
            break

    # Если пользователя нет, добавляем новую запись
    if not user_exists:
        sheet.append([user_name, mark, record_time_string])

    # Сохраняем изменения
    workbook.save(file_name)


def write_beck_to_file(file_name: str, user_name: str, beck_test_result: int) -> None:
    """Записывает информацию по результатам теста Бека в файл"""
    # Проверяем, существует ли файл
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        # Если файл не существует, создаем новый и добавляем заголовки
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["user_name", "beck_test_result"])  # Заголовки столбцов

    # Флаг для отслеживания наличия пользователя
    user_exists = False

    # Проверяем существующие записи
    for row in sheet.iter_rows(min_row=2):  # Пропускаем заголовок
        if row[0].value == user_name:
            row[1].value = beck_test_result  # Обновляем результат теста Бека для пользователя
            user_exists = True
            break

    # Если пользователя нет, добавляем новую запись
    if not user_exists:
        sheet.append([user_name, beck_test_result])

    # Сохраняем изменения в файле
    workbook.save(file_name)


def write_schedule_to_file(file_name: str, user_name: str, schedule: str) -> None:
    """Записывает информацию по расписанию в файл"""
    # Проверяем, существует ли файл
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        # Если файл не существует, создаем новый и добавляем заголовки
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["user_name", "schedule"])  # Заголовки столбцов

    # Флаг для отслеживания наличия пользователя
    user_exists = False

    # Проверяем существующие записи
    for row in sheet.iter_rows(min_row=2):  # Пропускаем заголовок
        if row[0].value == user_name:
            row[1].value = schedule  # Обновляем время для существующего пользователя
            user_exists = True
            break

    # Если пользователя нет, добавляем новую запись
    if not user_exists:
        sheet.append([user_name, schedule])

    # Сохраняем изменения в файле
    workbook.save(file_name)


def read_data_from_file(file_name: str, user_name: str) -> list:
    """Считывает историю записей из архива"""
    # Загружаем рабочую книгу
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    filtered_rows = []

    # Проходим по всем строкам в листе
    for row in sheet.iter_rows(values_only=True):
        # Предполагаем, что имя пользователя находится в первом столбце (индекс 0)
        if row[0] == user_name:
            filtered_rows.append(row)

    return filtered_rows


@bot.message_handler(commands=['start'])
def start(message) -> None:
    """Создаём кнопки телеграм-бота"""
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button1 = types.KeyboardButton('Записать настроение')
    button2 = types.KeyboardButton('Календарь настроения')
    button3 = types.KeyboardButton('Рекомендации')
    button4 = types.KeyboardButton('Авторизация')
    button5 = types.KeyboardButton('Пройти тест Бека')
    button6 = types.KeyboardButton('Время напоминания')
    markup.add(button1, button2, button3, button4, button5, button6)

    bot.send_message(message.chat.id, 'Привет! Я бот для отслеживания настроения.')
    bot.send_message(message.chat.id, 'Выберите кнопку:', reply_markup=markup)
    bot.register_next_step_handler(message, handle_message)


@bot.message_handler(func=lambda message: True)
def handle_message(message):
    """Реализуем работу всех кнопок"""
    if message.chat.id not in user_names and message.text == 'Авторизация':
        bot.send_message(message.chat.id, f'Для авторизации введите Ваше имя:')
        bot.register_next_step_handler(message, authorization)

    elif message.chat.id not in user_names and message.text != 'Авторизация':
        bot.send_message(message.chat.id, f'Для начала работы бота, пожалуйста, авторизуйтесь.')
        bot.register_next_step_handler(message, handle_message)

    elif message.text == 'Записать настроение':
        bot.send_message(message.chat.id, 'Оцените своё сегодняшнее настроение по шкале от 1 до 10.')
        bot.register_next_step_handler(message, write_mood)

    elif message.text == 'Календарь настроения':
        check_mood_calender(message)

    elif message.text == 'Рекомендации':
        repeat_recommendation(message)

    elif message.text == 'Пройти тест Бека':
        start_test(message)

    elif message.text == 'Время напоминания':
        bot.send_message(message.chat.id, 'Введите время, в которое бот будет каждый день напоминать Вам заходить и '
                                          'записывать своё настроение, в формате ЧЧ:ММ по Московскому времени.')
        bot.register_next_step_handler(message, set_reminder_time)

    elif message.text == 'Авторизация':
        bot.send_message(message.chat.id, 'Вы уже авторизованы!')

    else:
        bot.send_message(message.chat.id, 'Пожалуйста, выберите одну из кнопок.')


def authorization(message):
    """Авторизация пользователя"""
    user_name = message.text
    user_names[message.chat.id] = user_name  # Сохраняем имя пользователя
    bot.send_message(message.chat.id, f'Приятно познакомиться, {user_name}!')


def write_mood(message):
    """Записываем настроение и дату записи в файл"""
    mood = message.text

    if mood in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']:
        write_stat_to_file(file_name_stat, user_names[message.chat.id], int(mood),
                           datetime.datetime.now().strftime('%Y-%m-%d'))
        bot.send_message(message.chat.id, 'Спасибо! Я запомнил ваш выбор.')

        if calculate_depression_warning(create_mood_calendar(message)):
            # Проверка на низкий показатель ответов
            bot.send_message(message.chat.id, 'Я заметил, что в последнее время вы часто указываете'
                                              'плохое настроение. Рекомендую пройти Вам тест Бека'
                                              'на депрессию.')

    else:
        bot.send_message(message.chat.id, 'Пожалуйста, напишите число от 1 до 10.')
        bot.register_next_step_handler(message, write_mood)


def create_mood_calendar(message) -> dict:
    """Создаем календарь настроения"""
    data = read_data_from_file(file_name_stat, user_names[message.chat.id])
    available_dates = [i[2] for i in data]
    dates = []
    final_dates = []
    moods = [i[1] for i in data]
    for j in range(20, -1, -1):
        new_date = datetime.datetime.now() - datetime.timedelta(j)
        new_date = new_date.strftime('%Y-%m-%d')
        dates.append(new_date)

    k = 0
    for date in dates:
        if date in available_dates:
            final_dates.append(moods[-k - 1])
            k += 1
        else:
            final_dates.append('В этот день не было записей.')
    return dict(zip(dates, final_dates[::-1]))


def check_mood_calender(message):
    """Вывод календаря настроений пользователю"""
    bot.send_message(message.chat.id, convert_mood_calendar(create_mood_calendar(message)))


def repeat_recommendation(message):
    """Повтор последней рекомендации"""
    try:
        beck_test_score = read_data_from_file(file_name_beck, user_names[message.chat.id])[0][1]
        last_recommendation = give_recommendation(int(beck_test_score))
        bot.send_message(message.chat.id, last_recommendation)
    except IndexError:
        bot.send_message(message.chat.id, 'В данный момент для вас нету рекомендаций.')


def convert_mood_calendar(calendar: dict) -> str:
    """Переделываем календарь из словаря в строку для дальнейшей отправки пользователю"""
    res = ''
    for key, value in calendar.items():
        res += f'{key}: {value}\n'
    return res


def calculate_depression_warning(calender: dict) -> bool:
    """Рассчитываем предупреждение о депрессии"""
    moods_values = [int(value) for value in calender.values() if value != 'В этот день не было записей.']
    if len(moods_values) >= 14:
        moods_values = moods_values[-14:]
    return (sum(moods_values) / 14) <= 4 and len(moods_values) >= 14


def start_test(message):
    """Начало прохождения теста Бека"""
    bot.send_message(message.chat.id, 'Вы начали проходить тест Бека на депрессию! '
                                      'Пожалуйста, ответьте на следующие вопросы, '
                                      'выбрав один из четырех вариантов (0-3):')
    ask_question(message.chat.id, 0, 0)


def ask_question(chat_id, question_index, total_score):
    """Поочерёдно задаём пользователю вопрос и в конце выдаём рекомендацию"""
    if question_index < len(questions):
        bot.send_message(chat_id, questions[question_index])
        bot.register_next_step_handler_by_chat_id(chat_id, process_answer, question_index, total_score)

    else:
        write_beck_to_file(file_name_beck, user_names[chat_id], total_score)
        last_recommendation = give_recommendation(total_score)
        bot.send_message(chat_id, last_recommendation)


def process_answer(message, question_index, total_score):
    """Получаем ответ на вопрос и проверяем его на корректность"""
    try:
        answer = int(message.text)
        if answer < 0 or answer > 3:
            raise ValueError('Ответ должен быть от 0 до 3.')

        total_score += answer
        ask_question(message.chat.id, question_index + 1, total_score)

    except ValueError:
        bot.send_message(message.chat.id, 'Пожалуйста, введите число от 0 до 3.')
        ask_question(message.chat.id, question_index, total_score)


def give_recommendation(test_score: int) -> str:
    """Выдаём рекомендации по результатам теста Бека"""
    if 0 <= test_score <= 13:
        return ('Согласно тесту Бека, у Вас нет депрессии. Для поддержания ментального '
                'здоровья рекомендую Вам следующие фишки из самопомощи:')

    elif 14 <= test_score <= 19:
        return ('Согласно тесту Бека, у Вас легкая депрессия. Для поддержания ментального '
                'здоровья рекомендую Вам следующие фишки из самопомощи:')

    elif 20 <= test_score <= 28:
        return ('Согласно тесту Бека, у Вас умеренная депрессия. Рекомендую вам обратиться за помощью '
                'к психиатру. \nТелефон единой помощи: 8 800 222-55-71')

    elif 29 <= test_score <= 63:
        return ('Согласно тесту Бека, у Вас тяжёлая депрессия. Рекомендую вам обратиться за помощью '
                'к психиатру. \nТелефон единой помощи: 8 800 222-55-71')


def set_reminder_time(message):
    """Устанавливаем пользовательское время напоминания"""
    # Извлекаем время из сообщения
    time_str = message.text
    # Проверяем корректность формата времени
    hour, minute = map(int, time_str.split(':'))

    if 0 <= hour < 23 and 0 <= minute < 60:
        user_id = message.chat.id
        reminder_time = f'{hour:02}:{minute:02}'
        write_schedule_to_file(file_name_schedule, user_names[message.chat.id], reminder_time)
        schedule_reminder(user_id)
        bot.reply_to(message, f'Время напоминания установлено на {reminder_time}.')
    else:
        bot.reply_to(message, 'Пожалуйста, укажите корректное время в формате HH:MM.')


def send_reminder(user_id):
    """Отправляем напоминание пользователю"""
    bot.send_message(user_id, 'Напоминание: Не забудьте зайти в бота и указать своё настроение!')


def schedule_reminder(user_id):
    """Получаем время напоминания из файла и каждый день в это время отправляем напоминание пользователю"""
    reminder_time = read_data_from_file(file_name_schedule, user_names[user_id])[0][1]
    shl.every().day.at(reminder_time).do(send_reminder, user_id)


def schedule_reminders():
    while True:
        shl.run_pending()
        time.sleep(1)


# Запуск планировщика в отдельном потоке
threading.Thread(target=schedule_reminders).start()


if __name__ == '__main__':
    bot.polling(none_stop=True)
